import os
import argparse
import json
import re
import requests
import sys
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()
SOLUTIONS_DIR = "solutions"
TESTCASES_DIR = "testcases"
REPORT_FILE = "dsa_progress_report.xlsx"
api_key = os.getenv("GEMINI_API_KEY")
api_url = f"https://generativelanguage.googleapis.com/v1/models/gemini-2.5-flash:generateContent?key={api_key}"


def generate_filename_with_ai(problem_description):
    """
    Uses Gemini API to generate a concise, snake_cased filename for the problem.
    """
    print("🤖 Asking Gemini to suggest a filename...")
    prompt = f"""
    Given the following DSA problem description or URL, generate a concise, snake_cased filename suitable for a Python file.
    Only return the filename string, without any extension or extra text. Avoid special characters and keep it under 60 characters.

    Problem:
    ---
    {problem_description}
    ---
    """
    payload = {
        "contents": [{"parts": [{"text": prompt}]}]
    }
    headers = {'Content-Type': 'application/json'}
    try:
        response = requests.post(api_url, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        raw_text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
        filename = raw_text.strip().replace(".py", "")
        filename = re.sub(r'[^\w_]', '', filename)
        filename = filename[:60]
        print(f"📝 Gemini suggested filename: {filename}")
        return filename or "untitled_problem"
    except Exception as e:
        print(f"Error generating filename with AI: {e}", file=sys.stderr)
        return "untitled_problem"


def get_gemini_api_key():
    """Fetches the Gemini API key from environment variables."""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        print("Error: GEMINI_API_KEY environment variable not found.", file=sys.stderr)
        print("Please create a .env file with GEMINI_API_KEY='your_key_here' or set it in your shell.", file=sys.stderr)
        return None
    return api_key


def update_progress_report(topic, title, statement, file_path, url):
    """
    Creates or updates an Excel progress report with the new problem's details.
    """
    print(f"📊 Updating progress report: {REPORT_FILE}")

    headers = ["Topic", "Problem Title", "Problem Statement", "Solution File Path", "Problem URL", "Status"]

    if os.path.exists(REPORT_FILE):
        workbook = load_workbook(REPORT_FILE)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[3] == file_path:
            print("⚠️  Problem already exists in the report. Skipping.")
            return

    new_row_data = [topic, title, statement, file_path, url, "Not Done"]
    sheet.append(new_row_data)

    dv = DataValidation(type="list", formula1='"Not Done,Done,Revisit"', allow_blank=True)
    dv.add('F2:F1000')
    sheet.add_data_validation(dv)

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = min(adjusted_width, 70)

    workbook.save(REPORT_FILE)
    print("✅ Report updated successfully.")


def generate_assets_with_ai(problem_description_or_url):
    """
    Calls the Gemini API to generate solution boilerplate, test cases, category,
    and (if a URL is provided) extracts the problem title and description using LLM.
    Returns a dict with keys: solution_boilerplate, test_cases, category, title, description.
    """
    print(f"🤖 Contacting Gemini API with the problem description or URL...")

    title = None
    description = None
    if re.match(r'https?://', problem_description_or_url.strip()):
        extract_prompt = f"""
        Given the following coding problem URL, extract:
        1. The exact problem title
        2. The complete problem description/statement

        Return the information in JSON format with two fields:
        - "title": The problem title
        - "description": The complete problem description

        URL:
        {problem_description_or_url}
        """
        payload_extract = {
            "contents": [{"parts": [{"text": extract_prompt}]}]
        }
        headers = {'Content-Type': 'application/json'}
        try:
            response = requests.post(api_url, headers=headers, json=payload_extract, timeout=60)
            response.raise_for_status()
            data = response.json()
            raw_text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "{}")
            json_match = re.search(r'\{.*\}', raw_text, re.DOTALL)
            info = json.loads(json_match.group()) if json_match else {}
            title = info.get("title")
            description = info.get("description")
            print(f"🔎 Extracted title: {title}")
        except Exception as e:
            print(f"Error extracting title/description from URL: {e}", file=sys.stderr)

    main_problem_text = description or problem_description_or_url

    prompt = f"""
    Based on the following DSA problem description, generate three assets:

    Problem Description:
    ---
    {main_problem_text}
    ---

    1.  A Python boilerplate code string. This string must contain two parts:
        a. A function named 'solve' with descriptive, type-hinted parameter names derived from the problem. It must include the full problem description in a comment, a detailed docstring, and a 'pass' statement.
        b. A main execution block (`if __name__ == "__main__":`). This block MUST be a self-contained test runner. It should import `json` and `pathlib`, automatically find the corresponding test file in the `../../testcases/CATEGORY/` directory, load the JSON, run all tests against the `solve` function, and print a clear summary of which tests passed or failed. The code should be robust and handle file-not-found errors. **IMPORTANT: The test runner must now look inside a category sub-folder.**

    2.  A list of 5-7 diverse JSON test cases. Each test case should be an object with two keys: "input" and "expected_output". The "input" key's value must be an object whose keys match the parameter names of the 'solve' function. Include common and edge cases.

    3.  A primary category for the problem. This should be a single, lowercase, snake_cased string. Choose from a predefined list if possible: array, string, hash_table, linked_list, stack, queue, tree, graph, binary_search, two_pointers, recursion, dynamic_programming, general.

    Return a single, valid JSON object with three top-level keys: "solution_boilerplate", "test_cases", and "category".
    Add the all the test cases all the edge cases should be covered.
    The value for "category" must be the category string.
    The value for "test_cases" must be the JSON array of test case objects.
    The value for "solution_boilerplate" must be the Python code as a single string.
    Do not include any text or markdown formatting outside of the main JSON object. Start the response with ```json.
    """

    payload = {
        "contents": [{"parts": [{"text": prompt}]}]
    }

    headers = {'Content-Type': 'application/json'}

    try:
        response = requests.post(api_url, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
        data = response.json()
        raw_text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "{}")
        json_match = re.search(r'```json\s*([\s\S]*?)\s*```', raw_text)
        clean_text = json_match.group(1) if json_match else raw_text
        parsed_content = json.loads(clean_text)
        required_keys = ["solution_boilerplate", "test_cases", "category"]
        if not all(key in parsed_content for key in required_keys):
            raise ValueError("AI response is missing one or more required keys: " + ", ".join(required_keys))
        print("✅ Successfully received assets from Gemini.")
        parsed_content["title"] = title
        parsed_content["description"] = description or main_problem_text
        return parsed_content
    except requests.exceptions.RequestException as e:
        print(f"Error calling Gemini API: {e}", file=sys.stderr)
    except (json.JSONDecodeError, ValueError, KeyError) as e:
        print(f"Error parsing AI response: {e}", file=sys.stderr)
        print("Raw response:", response.text if 'response' in locals() else "No response", file=sys.stderr)
    return None


def sanitize_filename(name):
    """Converts a problem name into a valid, reasonably short filename."""
    s = name.lower().strip()
    s = re.sub(r'[\s-]+', '_', s)
    s = re.sub(r'[^\w_]', '', s)
    return s[:60]


def sanitize_category(name):
    """Sanitizes the category name to ensure it's a valid directory name."""
    s = str(name).lower().strip()
    s = re.sub(r'[\s]+', '_', s)
    s = re.sub(r'[^a-z0-9_]', '', s)
    return s or "general"


def create_files(problem_description, problem_url):
    """
    Main function to generate the directory structure, AI-generated files, and update the report.
    Uses LLM to generate the filename.
    """
    api_key = get_gemini_api_key()
    if not api_key:
        return

    generated_assets = generate_assets_with_ai(problem_description)
    if not generated_assets:
        print("❌ Failed to generate assets. Aborting.", file=sys.stderr)
        return

    filename_base = generate_filename_with_ai(problem_description)
    if not filename_base:
        print("Error: Could not create a valid filename from the problem description.", file=sys.stderr)
        return

    category = sanitize_category(generated_assets.get("category", "general"))
    print(f"🧠 AI classified this problem under category: '{category}'")

    solution_subdir = os.path.join(SOLUTIONS_DIR, category)
    testcase_subdir = os.path.join(TESTCASES_DIR, category)

    os.makedirs(solution_subdir, exist_ok=True)
    os.makedirs(testcase_subdir, exist_ok=True)
    print(f"\nDirectories '{solution_subdir}/' and '{testcase_subdir}/' are ready.")

    solution_path = os.path.join(solution_subdir, f"{filename_base}.py")
    testcase_path = os.path.join(testcase_subdir, f"{filename_base}.json")

    if not os.path.exists(solution_path):
        with open(solution_path, 'w', encoding='utf-8') as f:
            f.write(generated_assets["solution_boilerplate"])
        print(f"✅ Created AI-generated solution file: {solution_path}")
    else:
        print(f"⚠️  Solution file already exists, skipping: {solution_path}")

    if not os.path.exists(testcase_path):
        with open(testcase_path, 'w', encoding='utf-8') as f:
            json.dump(generated_assets["test_cases"], f, indent=4)
        print(f"✅ Created AI-generated test case file: {testcase_path}")
    else:
        print(f"⚠️  Test case file already exists, skipping: {testcase_path}")

    print("\n🚀 Setup complete! Happy coding!")
    print(f"   1. Review and implement your logic in '{solution_path}'.")
    print(f"   2. Run 'python {solution_path}' to check your work.")

    # Use extracted title/description if available, else fallback
    excel_title = generated_assets.get("title") or filename_base
    excel_statement = generated_assets.get("description") or problem_description

    update_progress_report(
        topic=category,
        title=excel_title,
        statement=excel_statement,
        file_path=solution_path,
        url=problem_url
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scaffold a DSA problem using AI. Provide a problem description or URL as an argument or pipe the full description via stdin.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        "problem_description",
        nargs='?',
        default=None,
        type=str,
        help="The problem description or URL.\nIf omitted, the script will read the full problem description from stdin."
    )
    parser.add_argument(
        "-u", "--url",
        type=str,
        default="",
        help="Optional: The URL of the problem statement (e.g., from LeetCode)."
    )
    args = parser.parse_args()

    if args.problem_description:
        full_description = args.problem_description
        create_files(full_description, args.url)
    elif not sys.stdin.isatty():
        print("Reading problem description from stdin... (Press Ctrl+D or Ctrl+Z then Enter to finish)")
        full_description = sys.stdin.read().strip()
        if not full_description:
            print("Error: Stdin was empty. Aborting.", file=sys.stderr)
            sys.exit(1)
        create_files(full_description, args.url)
    else:
        print("No problem description provided and no input from stdin. Please see usage below.", file=sys.stderr)
        parser.print_help()
